
from django.shortcuts import render, redirect
from .forms import *
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from user_management.models import User
from django.shortcuts import get_object_or_404
from user_management.decorators import check_permission

def user_login(request):
    if request.method == 'POST':
        print('request.POST', request.POST)
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('email')
            password = form.cleaned_data.get('password')
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')  # Redirect to a success page
            else:
                print('Invalid email or password')
                form.add_error(None, 'Invalid email or password')
        else:
            print('form', form.errors)
    else:
        form = LoginForm()

    context = {
        'form': form
    }
    return render(request, 'Auth/login.html', context)

from django.utils.timezone import now

def dashboard(request):
    records = Patient.objects.all()
    today = now().date()
 
    sheets_today = DailySheet.objects.filter(date=today)
    pc_lists_today = PCList.objects.filter(date=today)


    return render(request, 'dashboard.html',{'dashboard':'active','total_count': records.count(), 'sheets_count': sheets_today.count(), 'pc_lists_count': pc_lists_today.count()})


def user_logout(request):
    if request.user.is_authenticated:
        logout(request)
        return redirect('user_login')
    else:
        return redirect('user_login')



@login_required(login_url='/')
@check_permission('patient_create')
def patient_create(request):
    try:
        if request.method == "POST":
            form = PatientForm(request.POST)
            if form.is_valid():
                # Additional case number validation
                case_number = form.cleaned_data.get('case_number')
                if case_number and not (case_number.startswith('UM') or case_number.startswith('PC')):
                    form.add_error('case_number', "Case number must start with 'UM' or 'PC'.")
                else:
                    form.save()
                    return redirect('patient_list')
            # If form is invalid (including duplicate case number), it will show errors
        else:
            form = PatientForm()
        
        context = {
            'form': form,
            'screen_name': 'Patient'
        }
        return render(request, 'create.html', context)
    except Exception as error:
        return render(request, '500.html', {'error': error})

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
@login_required(login_url='/')
@check_permission('patient_view')
def patient_list(request):
    """
    Displays a filtered list of Patient records with optional filtering
    - Supports filtering by case number, name, gender, year, and date range
    - Scrollable table (no pagination)
    """
    records = Patient.objects.all()
    
    # Initialize filter form with GET parameters
    filter_form = PatientFilterForm(request.GET or None)
    
    if filter_form.is_valid():
        case_number = filter_form.cleaned_data.get('case_number')
        if case_number:
            records = records.filter(case_number__icontains=case_number)
        
        name = filter_form.cleaned_data.get('name')
        if name:
            records = records.filter(name__icontains=name)
        
        gender = filter_form.cleaned_data.get('gender')
        if gender:
            records = records.filter(gender=gender)
        
        year = filter_form.cleaned_data.get('year')
        if year:
            records = records.filter(created_at__year=year)
        
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        if date_from:
            records = records.filter(created_at__date__gte=date_from)
        if date_to:
            records = records.filter(created_at__date__lte=date_to)
    
    context = {
        'records': records,       # Full queryset
        'patients': records,      # Alias if needed
        'filter_form': filter_form,
        'screen_name': 'Patient',
        'total_count': records.count()
    }
    
    return render(request, 'patient_list.html', context)

    
@login_required(login_url='/')
@check_permission('patient_update')
def patient_update(request, pk):
    """
    Handles updating an existing Patient record.
    - Validates case number uniqueness (excluding current record)
    """
    try:
        patient = get_object_or_404(Patient, pk=pk)
        
        if request.method == "POST":
            form = PatientForm(request.POST, instance=patient)
            if form.is_valid():
                form.save()
                return redirect('patient_list')
        else:
            form = PatientForm(instance=patient)
        
        context = {
            'form': form,
            'screen_name': 'Patient'
        }
        return render(request, 'create.html', context)
    except Exception as error:
        return render(request, '500.html', {'error': error})


@login_required(login_url='/')
@check_permission('patient_delete')
def patient_delete(request, pk):
    """
    Handles deletion of a Patient record.
    """
    try:
        patient = get_object_or_404(Patient, pk=pk)
        patient.delete()
        return redirect('patient_list')
    except Exception as error:
        return render(request, '500.html', {'error': error})
import pandas as pd
from django.http import HttpResponse
from io import BytesIO


def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            df = pd.read_excel(file)

            # Normalize headers
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

            for _, row in df.iterrows():
                Patient.objects.update_or_create(
                    case_number=row.get('case_number'),
                    defaults={
                        'name': row.get('name'),
                        'age': row.get('age'),
                        'gender': row.get('gender'),
                        'chief_complaint': row.get('chief_complaint'),
                        'reference': row.get('reference'),
                        'contact': row.get('contact_number') or row.get('contact'),
                        'address': row.get('address')
                    }
                )
            return redirect('patient_list')
    else:
        form = UploadFileForm()
    return render(request, 'upload_excel.html', {'form': form})


def download_excel(request):
    patients = Patient.objects.all()
    data = {
        'Name': [p.name for p in patients],
        'Case Number': [p.case_number for p in patients],
        'Age': [p.age for p in patients],
        'Gender': [p.get_gender_display() for p in patients],
        'Chief Complaint': [p.chief_complaint for p in patients],
        'Reference': [p.reference for p in patients],
        'Contact': [p.contact for p in patients],
        'Address': [p.address for p in patients],
    }
    
    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=patients.xlsx'
    return response


from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import DailySheet
from .forms import DailySheetFilterForm


@login_required(login_url='/')
def daily_sheet_list(request):
    sheets = DailySheet.objects.all().order_by('-date')

    filter_form = DailySheetFilterForm(request.GET or None)

    if filter_form.is_valid():
        case_number = filter_form.cleaned_data.get('case_number')
        if case_number:
            sheets = sheets.filter(case_number__icontains=case_number)

        name = filter_form.cleaned_data.get('name')
        if name:
            sheets = sheets.filter(name__icontains=name)

        payment_status = filter_form.cleaned_data.get('payment_status')
        if payment_status:
            sheets = sheets.filter(payment_status=payment_status)

        year = filter_form.cleaned_data.get('year')
        if year:
            sheets = sheets.filter(created_at__year=year)

        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            sheets = sheets.filter(date__gte=date_from)

        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            sheets = sheets.filter(date__lte=date_to)

    context = {
        'sheets': sheets,
        'filter_form': filter_form,
        'total_count': sheets.count(),
        'screen_name': 'Daily Sheet'
    }
    return render(request, 'daily_sheet_list.html', context)


from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.utils import timezone
from django.contrib import messages
from .models import DailySheet, Patient
from .forms import DailySheetForm

from django.utils import timezone
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from .models import Patient, DailySheet
from .forms import DailySheetForm


def daily_sheet_create(request):
    if request.method == 'POST':
        form = DailySheetForm(request.POST)
        if form.is_valid():
            daily_sheet = form.save(commit=False)
            
            # Get current time
            current_time = timezone.localtime()
            
            # Auto fill in_time if case_number is provided and in_time is empty
            if daily_sheet.case_number and not daily_sheet.in_time:
                daily_sheet.in_time = current_time
                
            # Auto fill out_time if payment_status is provided and out_time is empty
            if daily_sheet.payment_status and not daily_sheet.out_time:
                daily_sheet.out_time = current_time
                
            daily_sheet.save()
            messages.success(request, 'Daily sheet entry created successfully!')
            return redirect('daily_sheet_list')
    else:
        form = DailySheetForm()
        # Set initial date to today
        form.fields['date'].initial = timezone.now().date()

    return render(request, 'daily_sheet_form.html', {'form': form})

def check_case_number(request):
    case_number = request.GET.get('case_number')

    try:
        patient = Patient.objects.get(case_number=case_number)
        
        # Check if case number starts with "UM"
        if not case_number.startswith('UM'):
            return JsonResponse({
                'exists': False,
                'error': 'Case number must start with "UM" for daily sheets. Please use the UM series.'
            })
            
        return JsonResponse({
            'exists': True,
            'patient_id': patient.id,
            'name': patient.name,
            'diagnosis': patient.diagnosis,
            'age': patient.age,
            'gender': patient.gender,
        })
    except Patient.DoesNotExist:
        return JsonResponse({'exists': False})

def patient_followups(request, patient_id):
    sheets = DailySheet.objects.filter(patient_id=patient_id).order_by('date')

    data = []
    for s in sheets:
        data.append({
            'date': s.date.strftime('%d-%m-%Y') if s.date else '',
            'feedback': s.feedback,
            'explain_treatment': s.explain_treatment

        })

    return JsonResponse({'data': data})
def daily_sheet_update(request, pk):
    """Update an existing daily sheet entry"""
    sheet = get_object_or_404(DailySheet, pk=pk)
    
    if request.method == 'POST':
        form = DailySheetForm(request.POST, instance=sheet)
        if form.is_valid():
            form.save()
            messages.success(request, 'Daily sheet entry updated successfully!')
            return redirect('daily_sheet_list')
    else:
        form = DailySheetForm(instance=sheet)
    
    return render(request, 'daily_sheet_form.html', {'form': form, 'action': 'Update', 'sheet': sheet})

def daily_sheet_delete(request, pk):
    """Delete a daily sheet entry"""
    sheet = get_object_or_404(DailySheet, pk=pk)
    
    if request.method == 'POST':
        sheet.delete()
        messages.success(request, 'Daily sheet entry deleted successfully!')
        return redirect('daily_sheet_list')
    
    return render(request, 'daily_sheet_confirm_delete.html', {'sheet': sheet})
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta

from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from .models import DailySheet


def daily_sheet_export(request):
    """Export daily sheets to Excel (Safe, IST-based, Production-ready)"""

    # Base queryset
    sheets = DailySheet.objects.all()

    # Filters
    filter_type = request.GET.get('filter_type', 'all')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')

    today = timezone.localdate()

    if filter_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        sheets = sheets.filter(date__range=[start_date, end_date])

    elif filter_type == 'month':
        sheets = sheets.filter(date__year=today.year, date__month=today.month)

    elif filter_type == 'year':
        sheets = sheets.filter(date__year=today.year)

    elif filter_type == 'custom' and custom_start and custom_end:
        sheets = sheets.filter(date__range=[custom_start, custom_end])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Sheets'

    # Headers
    headers = [
        'Date', 'Name', 'Case Number', 'Diagnosis',
        'Charge', 'Received', 'Payment Status',
        'Payment Type', 'Payment Frequency',
        'In Time', 'Out Time',
        'Treatment 1', 'Treatment 2', 'Treatment 3', 'Treatment 4',
        'Therapist 1', 'Therapist 2'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    for row_num, sheet in enumerate(sheets, start=2):

        # Date (safe)
        date_cell = ws.cell(row=row_num, column=1)
        date_cell.value = sheet.date if sheet.date else ''
        date_cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

        ws.cell(row=row_num, column=2).value = sheet.name or ''
        ws.cell(row=row_num, column=3).value = sheet.case_number or ''
        ws.cell(row=row_num, column=4).value = sheet.diagnosis or ''

        ws.cell(row=row_num, column=5).value = float(sheet.charge) if sheet.charge else 0
        ws.cell(row=row_num, column=6).value = float(sheet.received) if sheet.received else 0

        ws.cell(row=row_num, column=7).value = sheet.get_payment_status_display() if sheet.payment_status else ''
        ws.cell(row=row_num, column=8).value = sheet.get_payment_type_display() if sheet.payment_type else ''
        ws.cell(row=row_num, column=9).value = sheet.get_payment_frequency_display() if sheet.payment_frequency else ''

        ws.cell(row=row_num, column=10).value = sheet.in_time.strftime('%H:%M') if sheet.in_time else ''
        ws.cell(row=row_num, column=11).value = sheet.out_time.strftime('%H:%M') if sheet.out_time else ''

        ws.cell(row=row_num, column=12).value = sheet.treatment_1 or ''
        ws.cell(row=row_num, column=13).value = sheet.treatment_2 or ''
        ws.cell(row=row_num, column=14).value = sheet.treatment_3 or ''
        ws.cell(row=row_num, column=15).value = sheet.treatment_4 or ''

        ws.cell(row=row_num, column=16).value = sheet.get_therapist_1_display() if sheet.therapist_1 else ''
        ws.cell(row=row_num, column=17).value = sheet.get_therapist_2_display() if sheet.therapist_2 else ''

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"daily_sheets_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def daily_sheet_import(request):
    """Import daily sheets from Excel"""
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                # Map payment status
                payment_status_map = {
                    'Paid': 'paid',
                    'Partially Paid': 'partially_paid',
                    'Not Paid': 'not_paid'
                }
                
                payment_type_map = {
                    'QR': 'qr',
                    'Cash': 'cash'
                }
                
                payment_frequency_map = {
                    'Daily Basis': 'daily',
                    'Weekly Basis': 'weekly',
                    'Monthly Basis': 'monthly'
                }
                
                therapist_map = {
                    'Dr. Basidh': 'dr_basidh',
                    'Dr. Visitra': 'dr_visitra',
                    'Dr. Bharatiselvi': 'dr_bharatiselvi'
                }
                
                imported_count = 0
                
                for index, row in df.iterrows():
                    try:
                        sheet = DailySheet(
                            date=pd.to_datetime(row['Date'], dayfirst=True).date(),
                            name=row['Name'],
                            case_number=str(row['Case Number']),
                            diagnosis=row['Diagnosis'],
                            charge=float(row['Charge']),
                            received=float(row['Received']),
                            payment_status=payment_status_map.get(row['Payment Status'], 'not_paid'),
                            payment_type=payment_type_map.get(row['Payment Type'], 'cash'),
                            payment_frequency=payment_frequency_map.get(row['Payment Frequency'], 'daily'),
                            in_time=pd.to_datetime(row['In Time']).time() if pd.notna(row['In Time']) else None,
                            out_time=pd.to_datetime(row['Out Time']).time() if pd.notna(row['Out Time']) else None,
                            treatment_1=row.get('Treatment 1', ''),
                            treatment_2=row.get('Treatment 2', ''),
                            treatment_3=row.get('Treatment 3', ''),
                            treatment_4=row.get('Treatment 4', ''),
                            therapist_1=therapist_map.get(row['Therapist 1'], 'dr_basidh'),
                            therapist_2=therapist_map.get(row.get('Therapist 2', ''), None) if pd.notna(row.get('Therapist 2')) else None
                        )
                        sheet.save()
                        imported_count += 1
                    except Exception as e:
                        messages.warning(request, f'Error importing row {index + 2}: {str(e)}')
                        continue
                
                messages.success(request, f'Successfully imported {imported_count} records!')
                return redirect('daily_sheet_list')
                
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'daily_sheet_import.html', {'form': form})


@login_required(login_url='/')
def pc_list_list(request):
    pc_lists = PCList.objects.all().order_by('-date')

    filter_form = PCListFilterForm(request.GET or None)

    if filter_form.is_valid():
        case_number = filter_form.cleaned_data.get('case_number')
        if case_number:
            pc_lists = pc_lists.filter(case_number__icontains=case_number)

        name = filter_form.cleaned_data.get('name')
        if name:
            pc_lists = pc_lists.filter(name__icontains=name)

        payment_status = filter_form.cleaned_data.get('payment_status')
        if payment_status:
            pc_lists = pc_lists.filter(payment_status=payment_status)

        year = filter_form.cleaned_data.get('year')
        if year:
            pc_lists = pc_lists.filter(created_at__year=year)

        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            pc_lists = pc_lists.filter(date__gte=date_from)

        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            pc_lists = pc_lists.filter(date__lte=date_to)

    context = {
        'pc_lists': pc_lists,
        'filter_form': filter_form,
        'total_count': pc_lists.count(),
        'screen_name': 'PC List'
    }
    return render(request, 'pc_list_list.html', context)

def pc_list_create(request):
    if request.method == 'POST':
        form = PCListForm(request.POST)
        if form.is_valid():
            pc_list = form.save(commit=False)
            
            # Get current time
            current_time = timezone.localtime()
            
            # Auto fill in_time if case_number is provided and in_time is empty
            if pc_list.case_number and not pc_list.in_time:
                pc_list.in_time = current_time
                
            # Auto fill out_time if payment_status is provided and out_time is empty
            if pc_list.payment_status and not pc_list.out_time:
                pc_list.out_time = current_time
                
            pc_list.save()
            messages.success(request, 'PC List entry created successfully!')
            return redirect('pc_list_list')
    else:
        form = PCListForm()
        # Set initial date to today
        form.fields['date'].initial = timezone.now().date()

    return render(request, 'pc_list_form.html', {'form': form})

def check_pc_number(request):
    case_number = request.GET.get('case_number')

    try:
        patient = Patient.objects.get(case_number=case_number)
        
        # Check if case number starts with "PC"
        if not case_number.startswith('PC'):
            return JsonResponse({
                'exists': False,
                'error': 'Case number must start with "PC" for PC lists. Please use the PC series.'
            })
            
        return JsonResponse({
            'exists': True,
            'patient_id': patient.id,
            'name': patient.name,
            'diagnosis': patient.diagnosis,
            'age': patient.age,
            'gender': patient.gender,
        })
    except Patient.DoesNotExist:
        return JsonResponse({'exists': False})

def patient_pc_followups(request, patient_id):
    pc_lists = PCList.objects.filter(patient_id=patient_id).order_by('date')

    data = []
    for s in pc_lists:
        data.append({
            'date': s.date.strftime('%d-%m-%Y') if s.date else '',
            'feedback': s.feedback,
            'explain_treatment': s.explain_treatment
        })

    return JsonResponse({'data': data})

def pc_list_update(request, pk):
    """Update an existing PC List entry"""
    pc_list = get_object_or_404(PCList, pk=pk)
    
    if request.method == 'POST':
        form = PCListForm(request.POST, instance=pc_list)
        if form.is_valid():
            form.save()
            messages.success(request, 'PC List entry updated successfully!')
            return redirect('pc_list_list')
    else:
        form = PCListForm(instance=pc_list)
    
    return render(request, 'pc_list_form.html', {'form': form, 'action': 'Update', 'pc_list': pc_list})

def pc_list_delete(request, pk):
    """Delete a PC List entry"""
    pc_list = get_object_or_404(PCList, pk=pk)
    
    if request.method == 'POST':
        pc_list.delete()
        messages.success(request, 'PC List entry deleted successfully!')
        return redirect('pc_list_list')
    
    return render(request, 'pc_list_confirm_delete.html', {'pc_list': pc_list})

def pc_list_export(request):
    """Export PC Lists to Excel"""
    # Same logic as daily_sheet_export but using PCList model
    pc_lists = PCList.objects.all()
    
    # [Filtering logic identical to daily_sheet_export]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PC Lists'
    
    # [Headers and data rows logic identical to daily_sheet_export]
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"pc_lists_{timezone.localtime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def pc_list_import(request):
    """Import PC Lists from Excel"""
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                # [Mapping logic identical to daily_sheet_import]
                
                imported_count = 0
                
                for index, row in df.iterrows():
                    try:
                        pc_list = PCList(
                            date=pd.to_datetime(row['Date'], dayfirst=True).date(),
                            name=row['Name'],
                            case_number=str(row['Case Number']),
                            diagnosis=row['Diagnosis'],
                            charge=float(row['Charge']),
                            received=float(row['Received']),
                            payment_status=payment_status_map.get(row['Payment Status'], 'not_paid'),
                            payment_type=payment_type_map.get(row['Payment Type'], 'cash'),
                            payment_frequency=payment_frequency_map.get(row['Payment Frequency'], 'daily'),
                            in_time=pd.to_datetime(row['In Time']).time() if pd.notna(row['In Time']) else None,
                            out_time=pd.to_datetime(row['Out Time']).time() if pd.notna(row['Out Time']) else None,
                            treatment_1=row.get('Treatment 1', ''),
                            treatment_2=row.get('Treatment 2', ''),
                            treatment_3=row.get('Treatment 3', ''),
                            treatment_4=row.get('Treatment 4', ''),
                            therapist_1=therapist_map.get(row['Therapist 1'], 'Basidh'),
                            therapist_2=therapist_map.get(row.get('Therapist 2', ''), None) if pd.notna(row.get('Therapist 2')) else None
                        )
                        pc_list.save()
                        imported_count += 1
                    except Exception as e:
                        messages.warning(request, f'Error importing row {index + 2}: {str(e)}')
                        continue
                
                messages.success(request, f'Successfully imported {imported_count} records!')
                return redirect('pc_list_list')
                
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'pc_list_import.html', {'form': form})


from collections import defaultdict
from django.db.models import Sum


def patient_ledger(patient):
    sheets = DailySheet.objects.filter(patient_id=patient).order_by("date", "id")
    balance = 0
    ledger = []

    for s in sheets:
        charge = s.charge or 0
        received = s.received or 0
        balance += received - charge

        ledger.append({
            "date": s.date,
            "case_number": s.case_number,
            "contact": s.patient_id.contact,
            "charge": charge,
            "received": received,
            "balance": balance,
        })
    return ledger, balance

def payment_dashboard(request):
    total_charge = DailySheet.objects.aggregate(t=Sum("charge"))["t"] or 0
    total_received = DailySheet.objects.aggregate(t=Sum("received"))["t"] or 0

    balance = total_received - total_charge

    context = {
        "total_charge": total_charge,
        "total_received": total_received,
        "total_pending": abs(balance) if balance < 0 else 0,
        "total_advance": balance if balance > 0 else 0,
    }
    return render(request, "accounts/payment_dashboard.html", context)

def pending_list(request):
    data = []

    for patient in Patient.objects.all():
        _, balance = patient_ledger(patient)
        if balance < 0:
            data.append({
                "patient": patient,
                "pending": abs(balance)
            })

    return render(request, "accounts/pending_list.html", {"data": data})

def advance_list(request):
    data = []

    for patient in Patient.objects.all():
        _, balance = patient_ledger(patient)
        if balance > 0:
            data.append({
                "patient": patient,
                "advance": balance
            })

    return render(request, "accounts/advance_list.html", {"data": data})


def patient_ledger_view(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    ledger, balance = patient_ledger(patient)

    return render(request, "accounts/patient_ledger.html", {
        "patient": patient,
        "ledger": ledger,
        "final_balance": balance
    })


def monthly_summary(request):
    records = DailySheet.objects.order_by("date")

    summary = defaultdict(lambda: {
        "charge": 0, "received": 0, "opening": 0, "closing": 0
    })

    balance = 0

    for r in records:
        key = r.date.strftime("%Y-%m")
        charge = r.charge or 0
        received = r.received or 0

        if summary[key]["charge"] == 0 and summary[key]["received"] == 0:
            summary[key]["opening"] = balance

        summary[key]["charge"] += charge
        summary[key]["received"] += received
        balance += received - charge
        summary[key]["closing"] = balance

    return render(request, "accounts/monthly_summary.html", {"summary": dict(summary)})

def yearly_summary(request):
    records = DailySheet.objects.order_by("date")

    summary = defaultdict(lambda: {
        "charge": 0, "received": 0, "opening": 0, "closing": 0
    })

    balance = 0

    for r in records:
        key = r.date.year
        charge = r.charge or 0
        received = r.received or 0

        if summary[key]["charge"] == 0 and summary[key]["received"] == 0:
            summary[key]["opening"] = balance

        summary[key]["charge"] += charge
        summary[key]["received"] += received
        balance += received - charge
        summary[key]["closing"] = balance

    return render(request, "accounts/yearly_summary.html", {"summary": dict(summary)})

from collections import defaultdict
from django.db.models import Sum
from django.shortcuts import render, get_object_or_404

def pc_patient_ledger_calc(patient):
    records = PCList.objects.filter(patient_id=patient).order_by("date", "id")
    balance = 0
    ledger = []

    for r in records:
        charge = r.charge or 0
        received = r.received or 0
        balance += received - charge

        ledger.append({
            "date": r.date,
            "case_number": r.case_number,
            "charge": charge,
            "received": received,
            "balance": balance,
        })

    return ledger, balance
def pc_payment_dashboard(request):
    total_charge = PCList.objects.aggregate(t=Sum("charge"))["t"] or 0
    total_received = PCList.objects.aggregate(t=Sum("received"))["t"] or 0

    balance = total_received - total_charge

    context = {
        "total_charge": total_charge,
        "total_received": total_received,
        "total_pending": abs(balance) if balance < 0 else 0,
        "total_advance": balance if balance > 0 else 0,
    }
    return render(request, "accounts/pc/payment_dashboard.html", context)

def pc_pending_list(request):
    data = []

    for patient in Patient.objects.all():
        _, balance = pc_patient_ledger_calc(patient)
        if balance < 0:
            data.append({
                "patient": patient,
                "pending": abs(balance)
            })

    return render(request, "accounts/pc/pending_list.html", {"data": data})
def pc_advance_list(request):
    data = []

    for patient in Patient.objects.all():
        _, balance = pc_patient_ledger_calc(patient)
        if balance > 0:
            data.append({
                "patient": patient,
                "advance": balance
            })

    return render(request, "accounts/pc/advance_list.html", {"data": data})
def pc_patient_ledger(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    ledger, balance = pc_patient_ledger_calc(patient)

    return render(request, "accounts/pc/patient_ledger.html", {
        "patient": patient,
        "ledger": ledger,
        "final_balance": balance
    })
def pc_monthly_summary(request):
    records = PCList.objects.order_by("date")

    summary = defaultdict(lambda: {
        "charge": 0, "received": 0, "opening": 0, "closing": 0
    })

    balance = 0

    for r in records:
        key = r.date.strftime("%Y-%m")
        charge = r.charge or 0
        received = r.received or 0

        if summary[key]["charge"] == 0 and summary[key]["received"] == 0:
            summary[key]["opening"] = balance

        summary[key]["charge"] += charge
        summary[key]["received"] += received
        balance += received - charge
        summary[key]["closing"] = balance

    return render(request, "accounts/pc/monthly_summary.html", {"summary": dict(summary)})
def pc_yearly_summary(request):
    records = PCList.objects.order_by("date")

    summary = defaultdict(lambda: {
        "charge": 0, "received": 0, "opening": 0, "closing": 0
    })

    balance = 0

    for r in records:
        key = r.date.year
        charge = r.charge or 0
        received = r.received or 0

        if summary[key]["charge"] == 0 and summary[key]["received"] == 0:
            summary[key]["opening"] = balance

        summary[key]["charge"] += charge
        summary[key]["received"] += received
        balance += received - charge
        summary[key]["closing"] = balance

    return render(request, "accounts/pc/yearly_summary.html", {"summary": dict(summary)})
